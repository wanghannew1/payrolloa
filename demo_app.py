import os
import re
import json
import time
from datetime import datetime
from io import BytesIO

import requests
import openpyxl
import streamlit as st
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

APP_KEY = os.getenv("DINGTALK_APP_KEY", "")
APP_SECRET = os.getenv("DINGTALK_APP_SECRET", "")
AGENT_ID = int(os.getenv("DINGTALK_AGENT_ID", "0") or "0")
PROCESS_CODE = os.getenv("DINGTALK_PROCESS_CODE", "")

DEFAULT_CONFIG = {
    "excel": {
        "title_row": 1,           # 报表名所在行（1-based）
        "unit_name_row": 2,       # 单位名所在行（1-based）；0 表示无此行，单位名从标题行用正则提取
        "header_start_row": 3,    # 表头起始行（1-based）
        "header_row_count": 3,    # 表头占几行（适配多级合并表头）
        "summary_row_marker": "合计",
        "unit_name_patterns": [
            r"(?:单位名称[：:]|[名称][：:]\s*)(.+?)(?:\s|$)"
        ],
        # 从标题行提取单位名（unit_name_row=0 时启用）
        # 算法：枚举标题里每个后缀的所有出现位置，从该位置向左扩展（只接受 ALLOWED 里的字符），
        # 得到所有候选后取「最长」一个。能正确处理「...有限公司净月团餐—分公司」这种嵌套后缀。
        # title_unit_patterns 是高级正则逃生口：若配置非空，则按列表里第一个能 search 命中的 group(1) 直接返回。
        "title_unit_suffixes": [
            "有限公司", "股份公司", "分公司", "公司", "集团",
            "医院", "卫生院", "诊所",
            "研究院", "研究所", "学院", "大学", "学校",
            "中心", "管委会", "事业部", "处", "局"
        ],
        "title_unit_allowed_chars": r"[一-龥A-Za-z0-9（）()·\-—]",
        "title_unit_patterns": [],
        "columns": {
            "transfer_total": {
                "keywords": ["转账合计"],
                "label": "转账合计（元）"
            },
            "deduction_total": {
                "keywords": ["扣款合计", "扣款"],
                "label": "扣款合计（五险一金、单位代理费）"
            },
            "net_total": {
                "keywords": ["实发合计", "实发工资", "实发"],
                "label": "实发合计（元）"
            },
            "personal_tax": {
                "keywords": ["个税", "个人所得税"],
                "label": "个人所得税"
            },
            "adjustment": {
                "keywords": ["调差", "差额调整", "调整差额", "工伤差额", "返还差额"],
                "label": "调差"
            },
            "service_fee": {
                "keywords": ["服务费", "代理费", "管理费"],
                "label": "服务费"
            },
            "employer_insurance": {
                "keywords": ["单位缴纳", "单位社保", "单位五险一金"],
                "label": "单位缴纳"
            }
        }
    },
    "validation": {
        "enabled": False,
        "strict": True,
        "tolerance": 0.00,
        "write_back_sheet": True,
        "write_back_sheet_name": "验证结果",
        "column_sum_checks": [],
        "row_formulas": []
    },
    "table_field": {
        "columns": [
            {"key": "report_name", "label": "报表名称"},
            {"key": "unit_name", "label": "甲方单位项目名称"},
            {"key": "transfer_total", "label": "转账合计（元）"},
            {"key": "deduction_total", "label": "扣款合计（五险一金、单位代理费）"},
            {"key": "net_total", "label": "实发合计（元）"},
            {"key": "tax_and_others", "label": "个人所得税及其他"}
        ]
    },
    "ui": {
        "template_name": "工资发放审批",
        "description": "请上传 Excel 工资表，系统将自动解析数据并提交钉钉 OA 审批流程。"
    }
}


def load_config():
    config_path = os.path.join(os.path.dirname(__file__), "config.json")
    if not os.path.exists(config_path):
        return DEFAULT_CONFIG
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return DEFAULT_CONFIG


CONFIG = load_config()


class DingTalkClient:
    def __init__(self, app_key, app_secret, agent_id, process_code):
        self.app_key = app_key
        self.app_secret = app_secret
        self.agent_id = agent_id
        self.process_code = process_code
        self.new_token = None  # v1.0 API token
        self.old_token = None  # oapi token
        self.token_expires = 0

    def _ensure_tokens(self):
        """Auto-refresh tokens if expired or not set."""
        now = time.time()
        if self.new_token and self.old_token and now < (self.token_expires - 300):
            return

        # New token (v1.0)
        resp = requests.post(
            "https://api.dingtalk.com/v1.0/oauth2/accessToken",
            json={"appKey": self.app_key, "appSecret": self.app_secret},
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        self.new_token = data["accessToken"]
        expire_in = data.get("expireIn", 7200)

        # Old token (oapi)
        resp = requests.get(
            "https://oapi.dingtalk.com/gettoken",
            params={"appkey": self.app_key, "appsecret": self.app_secret},
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        self.old_token = data["access_token"]
        old_expire = data.get("expires_in", 7200)

        self.token_expires = now + min(expire_in, old_expire)

    def get_user_by_mobile(self, mobile) -> str:
        """Return userId for a given mobile number."""
        self._ensure_tokens()
        resp = requests.post(
            f"https://oapi.dingtalk.com/topapi/v2/user/getbymobile?access_token={self.old_token}",
            json={"mobile": mobile},
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        if data.get("errcode") != 0:
            raise Exception(f"getbymobile error: {data}")
        return data["result"]["userid"]

    def get_user_info(self, user_id) -> dict:
        """Return {unionId, deptId, name} for a given userId."""
        self._ensure_tokens()
        resp = requests.post(
            f"https://oapi.dingtalk.com/topapi/v2/user/get?access_token={self.old_token}",
            json={"userid": user_id, "language": "zh_CN"},
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        if data.get("errcode") != 0:
            raise Exception(f"get user info error: {data}")
        result = data["result"]
        dept_list = result.get("dept_id_list", [])
        dept_id = dept_list[0] if dept_list else 0
        return {
            "unionId": result.get("unionid", ""),
            "deptId": dept_id,
            "name": result.get("name", ""),
        }

    def authorize_upload(self, user_id) -> str:
        """Return spaceId. Call BEFORE each upload to grant temporary permission."""
        self._ensure_tokens()
        resp = requests.post(
            "https://api.dingtalk.com/v1.0/workflow/processInstances/spaces/infos/query",
            headers={
                "x-acs-dingtalk-access-token": self.new_token,
                "Content-Type": "application/json",
            },
            json={"userId": user_id, "agentId": self.agent_id},
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        return str(data["result"]["spaceId"])

    def upload_file(self, file_bytes, file_name, union_id, space_id) -> dict:
        """Upload file to DingTalk storage and return {fileId, fileName, fileSize}."""
        self._ensure_tokens()
        file_size = len(file_bytes)
        headers = {
            "x-acs-dingtalk-access-token": self.new_token,
            "Content-Type": "application/json",
        }

        # Step 1: query upload info
        resp = requests.post(
            f"https://api.dingtalk.com/v1.0/storage/spaces/{space_id}/files/uploadInfos/query",
            headers=headers,
            params={"unionId": union_id},
            json={
                "protocol": "HEADER_SIGNATURE",
                "multipart": False,
                "option": {
                    "storageDriver": "DINGTALK",
                    "preCheckParam": {
                        "size": file_size,
                        "parentId": "0",
                        "name": file_name,
                    },
                },
            },
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        upload_key = data["uploadKey"]
        resource_url = data["headerSignatureInfo"]["resourceUrls"][0]
        oss_headers = data["headerSignatureInfo"]["headers"]

        # Step 2: PUT file to OSS
        upload_headers = dict(oss_headers)
        upload_headers["Content-Type"] = ""  # MUST be empty string
        resp = requests.put(
            resource_url,
            data=file_bytes,
            headers=upload_headers,
            timeout=60,
        )
        if resp.status_code != 200:
            raise Exception(f"OSS upload failed: {resp.status_code}")

        # Step 3: commit
        resp = requests.post(
            f"https://api.dingtalk.com/v1.0/storage/spaces/{space_id}/files/commit",
            headers=headers,
            params={"unionId": union_id},
            json={
                "uploadKey": upload_key,
                "name": file_name,
                "parentId": "0",
                "option": {
                    "size": file_size,
                    "conflictStrategy": "AUTO_RENAME",
                },
            },
            timeout=30,
        )
        resp.raise_for_status()
        dentry = resp.json()["dentry"]
        return {
            "fileId": str(dentry["id"]),
            "fileName": dentry["name"],
            "fileSize": dentry["size"],
            "spaceId": space_id,
            "fileType": dentry.get("extension", "xlsx"),
        }

    def create_approval(
        self, user_id, dept_id, title, attachments, table_rows, note=""
    ) -> str:
        """Create approval instance and return instanceId."""
        self._ensure_tokens()
        headers = {
            "x-acs-dingtalk-access-token": self.new_token,
            "Content-Type": "application/json",
        }

        attachment_value = json.dumps(
            [
                {
                    "spaceId": a["spaceId"],
                    "fileName": a["fileName"],
                    "fileSize": a["fileSize"],
                    "fileType": a["fileType"],
                    "fileId": a["fileId"],
                }
                for a in attachments
            ],
            ensure_ascii=False,
        )

        table_value = json.dumps(table_rows, ensure_ascii=False)

        form_values = [
            {"name": "标题", "value": title},
            {"name": "批量上传工资表", "value": attachment_value},
            {"name": "表格", "value": table_value},
            {"name": "备注", "value": note},
        ]

        resp = requests.post(
            "https://api.dingtalk.com/v1.0/workflow/processInstances",
            headers=headers,
            json={
                "processCode": self.process_code,
                "originatorUserId": user_id,
                "deptId": dept_id,
                "formComponentValues": form_values,
                "title": title,
            },
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        return data.get("instanceId", "")


def _extract_year_month(text, current_year=None):
    """
    从一段文本里抽取「YYYY年MM月」并归一化为 4位年+2位月 字符串。

    支持的形态（按从严到宽顺序）：
      1) 4位年+1或2位月，连体：  2026年5月  / 2026年05月
      2) 2位年+1或2位月，连体：  26年5月
      3) 分离形态：先找最右一个「YYYY年」或「YY年」，再在其后找最近的「N月」
         例：「...2026年派遣员工5月工资明细表」

    2 位年归一化规则：>= (current_year-50) 的两位数算 21 世纪，否则算 20 世纪。
    匹配不到返回 ""。
    """
    if not text:
        return ""
    if current_year is None:
        current_year = datetime.now().year

    def normalize(y, mo):
        y = int(y); mo = int(mo)
        if y < 100:
            # 2 位年补全：在 [current-50, current+49] 范围里选
            century_base = (current_year // 100) * 100  # 2000
            cand_new = century_base + y                 # 2026
            cand_old = cand_new - 100                   # 1926
            # 选离当前年最近、且距离 < 50 的那个
            if abs(cand_new - current_year) < 50:
                y = cand_new
            else:
                y = cand_old
        if not (1 <= mo <= 12):
            return ""
        return f"{y:04d}年{mo:02d}月"

    # 规则 1：4 位年连体
    m = re.search(r"(\d{4})年(\d{1,2})月", text)
    if m:
        out = normalize(m.group(1), m.group(2))
        if out:
            return out

    # 规则 2：2 位年连体（注意要求年前面不是数字，避免吃掉 4 位年的尾巴）
    m = re.search(r"(?<!\d)(\d{2})年(\d{1,2})月", text)
    if m:
        out = normalize(m.group(1), m.group(2))
        if out:
            return out

    # 规则 3：分离形态——先找最后一个 N年，再在其后找 N月
    year_match = None
    for ym in re.finditer(r"(?<!\d)(\d{4}|\d{2})年", text):
        year_match = ym
    if year_match:
        tail = text[year_match.end():]
        mo_match = re.search(r"(\d{1,2})月", tail)
        if mo_match:
            out = normalize(year_match.group(1), mo_match.group(1))
            if out:
                return out
    return ""


def parse_excel(file_bytes, filename):
    """
    Extract payroll data from Excel bytes.
    Returns dict with report_name, unit_name, year_month,
    transfer_total, deduction_total, net_total, tax_and_others.
    """
    # 根据扩展名选择解析库：.xlsx → openpyxl, .xls → xlrd
    filename_lower = filename.lower()
    if filename_lower.endswith('.xls') and not filename_lower.endswith('.xlsx'):
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes)
        ws = wb.sheet_by_index(0)
        rows = [ws.row_values(r) for r in range(ws.nrows)]
    else:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None

    excel_cfg = CONFIG.get("excel", DEFAULT_CONFIG["excel"])
    default_excel = DEFAULT_CONFIG["excel"]
    title_row_idx = int(excel_cfg.get("title_row", default_excel["title_row"])) - 1
    unit_name_row_cfg = int(excel_cfg.get("unit_name_row", default_excel["unit_name_row"]))
    header_start_idx = int(excel_cfg.get("header_start_row", default_excel["header_start_row"])) - 1
    header_row_count = int(excel_cfg.get("header_row_count", default_excel["header_row_count"]))

    # Title row: report name (first non-empty cell)
    report_name = ""
    if 0 <= title_row_idx < len(rows):
        for cell in rows[title_row_idx]:
            if cell is not None and str(cell).strip():
                report_name = str(cell).strip()
                break

    patterns = excel_cfg.get("unit_name_patterns", default_excel["unit_name_patterns"])
    title_patterns = excel_cfg.get("title_unit_patterns", default_excel["title_unit_patterns"])
    title_suffixes = excel_cfg.get("title_unit_suffixes", default_excel["title_unit_suffixes"])
    title_allowed = excel_cfg.get("title_unit_allowed_chars", default_excel["title_unit_allowed_chars"])

    # Unit name: from unit_name_row if configured, else extract from title row
    unit_name = ""
    if unit_name_row_cfg > 0:
        unit_row_idx = unit_name_row_cfg - 1
        if 0 <= unit_row_idx < len(rows):
            row_text = " ".join(
                str(cell).strip() for cell in rows[unit_row_idx] if cell is not None
            )
            matched = False
            for pattern in patterns:
                m = re.search(pattern, row_text)
                if m:
                    unit_name = m.group(1).strip()
                    matched = True
                    break
            if not matched:
                # Fallback: use first non-empty cell in that row
                for cell in rows[unit_row_idx]:
                    if cell is not None and str(cell).strip():
                        val = str(cell).strip()
                        if "名称" in val or "单位" in val:
                            unit_name = val.replace("单位", "").replace("名称", "").replace("：", "").replace(":", "").strip()
                            break
    else:
        # No dedicated unit row → extract from title row
        # 1) 高级用户可通过 title_unit_patterns 提供自定义正则（取第一个命中的 group 1）
        if title_patterns:
            for pattern in title_patterns:
                try:
                    m = re.search(pattern, report_name)
                except re.error:
                    continue
                if m and m.lastindex:
                    unit_name = m.group(1).strip()
                    break
        # 2) 否则用「后缀枚举 + 向左贪婪扩展 + 取最长」算法
        if not unit_name and title_suffixes:
            try:
                allowed_re = re.compile(title_allowed)
            except re.error:
                allowed_re = None
            if allowed_re is not None:
                candidates = []
                for suf in title_suffixes:
                    for m in re.finditer(re.escape(suf), report_name):
                        start = m.start()
                        while start > 0 and allowed_re.fullmatch(report_name[start - 1]):
                            start -= 1
                        cand = report_name[start : m.end()].strip()
                        if cand:
                            candidates.append(cand)
                if candidates:
                    unit_name = max(candidates, key=len)

    # Year month: 优先取标题（审计权威来源），文件名作为兜底
    # 同时单独保留两路结果，供 UI 做「标题 vs 文件名」一致性提醒
    year_month_from_title = _extract_year_month(report_name)
    year_month_from_filename = _extract_year_month(filename)
    year_month = year_month_from_title or year_month_from_filename

    summary_marker = excel_cfg.get("summary_row_marker", default_excel["summary_row_marker"])
    # Strip ALL whitespace (incl. internal) to tolerate variants like "合 计" / " 合计 "
    marker_normalized = re.sub(r"\s+", "", summary_marker)
    summary_row = None
    summary_row_idx = -1
    for ridx, row in enumerate(rows):
        if row and row[0] is not None:
            first_cell = re.sub(r"\s+", "", str(row[0]))
            if first_cell == marker_normalized:
                summary_row = row
                summary_row_idx = ridx
                break

    if summary_row is None:
        fallback = {
            "report_name": report_name,
            "unit_name": unit_name,
            "year_month": year_month,
            "year_month_from_title": year_month_from_title,
            "year_month_from_filename": year_month_from_filename,
            "transfer_total": "0.00",
            "deduction_total": "0.00",
            "net_total": "0.00",
            "tax_and_others": "0.00",
            "column_indices": {},
            "summary_row": None,
            "data_rows": [],
        }
        # 兜底也要把 excel.columns 里所有列填上 "0.00"，避免下游 KeyError
        for col_key in excel_cfg.get("columns", {}):
            fallback.setdefault(col_key, "0.00")
        return fallback

    # Header rows from config (1-based start, N rows)
    header_rows = rows[header_start_idx : header_start_idx + header_row_count]

    def find_col_index(keywords, exact_first=True):
        """Find column index matching keywords in header rows."""
        # Try exact match first
        for ridx, hrow in enumerate(header_rows):
            for cidx, cell in enumerate(hrow):
                if cell is None:
                    continue
                text = str(cell).strip()
                if not text:
                    continue
                for kw in keywords:
                    if kw in text:
                        return cidx
        return -1

    excel_cols = excel_cfg["columns"]

    # 通用化：把 excel.columns 里定义的每一列都找出索引，存到 column_indices
    # 同时按 keywords 顺序匹配（早出现的优先），跟原来 net_total 的逻辑一致
    column_indices = {}
    for col_key, col_def in excel_cols.items():
        keywords = col_def.get("keywords", [])
        idx = -1
        for kw in keywords:
            idx = find_col_index([kw])
            if idx != -1:
                break
        column_indices[col_key] = idx

    transfer_idx = column_indices.get("transfer_total", -1)
    deduction_idx = column_indices.get("deduction_total", -1)
    net_idx = column_indices.get("net_total", -1)

    def get_val(idx):
        if idx >= 0 and idx < len(summary_row):
            v = summary_row[idx]
            if v is None:
                return "0.00"
            try:
                return f"{float(v):.2f}"
            except (ValueError, TypeError):
                return str(v).strip() or "0.00"
        return "0.00"

    transfer_total = get_val(transfer_idx)
    deduction_total = get_val(deduction_idx)
    net_total = get_val(net_idx)

    # 把 excel.columns 里所有定义过的列，都从合计行取值，统一加到返回 dict
    # 这样 table_field 可以引用任意 excel.columns 里定义的 key（如 personal_tax, adjustment）
    column_summary_values = {}
    for col_key, cidx in column_indices.items():
        column_summary_values[col_key] = get_val(cidx)

    # tax_and_others 保留向后兼容（旧 config 可能还在用）
    try:
        tax_val = float(transfer_total) - float(deduction_total) - float(net_total)
        if abs(tax_val) < 0.005:
            tax_val = 0.0
        tax_and_others = f"{tax_val:.2f}"
    except (ValueError, TypeError):
        tax_and_others = "0.00"

    # 数据行 = 表头之后 到 合计行之前；过滤掉全空行
    header_end_idx = header_start_idx + header_row_count
    data_rows = []
    for r in rows[header_end_idx:summary_row_idx]:
        if r is None:
            continue
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in r):
            continue
        data_rows.append(r)

    result = {
        "report_name": report_name,
        "unit_name": unit_name,
        "year_month": year_month,
        "year_month_from_title": year_month_from_title,
        "year_month_from_filename": year_month_from_filename,
        "tax_and_others": tax_and_others,
        "column_indices": column_indices,
        "summary_row": summary_row,
        "data_rows": data_rows,
    }
    # 把 excel.columns 里所有列的合计值都加上（包括 transfer/deduction/net）
    # 这样 table_field 可以引用任意已配置的列
    result.update(column_summary_values)
    return result


def _to_money(v):
    """把单元格值转成 round 到 2 位小数的 float；None/空/非数 → 0.0。"""
    if v is None:
        return 0.0
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return 0.0
        try:
            return round(float(s), 2)
        except ValueError:
            return 0.0
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return 0.0


def _row_label(row, default_label):
    """数据行标识：序号+姓名 → 'row 8 张吉'；否则只用 default_label。"""
    if not row:
        return default_label
    first = row[0]
    name = row[1] if len(row) > 1 else None
    if isinstance(first, (int, float)) and name and str(name).strip():
        return f"{default_label} {str(name).strip()}"
    return default_label


def validate_payroll(parsed, validation_cfg, excel_cols):
    """
    校验解析后的工资表数据，返回结构化结果。
    parsed: parse_excel 的返回值（必须含 column_indices/summary_row/data_rows）
    validation_cfg: CONFIG['validation']
    excel_cols: CONFIG['excel']['columns']

    校验类型：
      column_sum:           sum(数据列) == 合计行该列
      row_formula_summary:  合计行 lhs == sum(rhs_plus) - sum(rhs_minus)
      row_formula_rows:     每个数据行同公式
    """
    tolerance = float(validation_cfg.get("tolerance", 0.0))
    column_indices = parsed.get("column_indices", {})
    summary_row = parsed.get("summary_row")
    data_rows = parsed.get("data_rows", [])

    checks = []

    def col_label(key):
        if key in excel_cols and excel_cols[key].get("label"):
            return excel_cols[key]["label"]
        return key

    def col_value(row, key):
        """从 row 中按 column_indices 取某 key 列的值（→ float）。"""
        idx = column_indices.get(key, -1)
        if idx < 0 or row is None or idx >= len(row):
            return 0.0
        return _to_money(row[idx])

    def require_col(key, where):
        if key not in excel_cols:
            raise ValueError(
                f"validation 配置中 {where} 引用了 '{key}'，但 excel.columns 未定义该列"
            )
        if column_indices.get(key, -1) < 0:
            # 列已声明但未在 Excel 表头中找到——这是数据/配置不匹配，不阻断校验
            # 取值会按 0 计，会触发对应的失败 issue（用户能从结果看出来）
            pass

    # === A. 纵向列加总 ===
    for spec in validation_cfg.get("column_sum_checks", []) or []:
        col_key = spec.get("column")
        if not col_key:
            continue
        require_col(col_key, "column_sum_checks")
        col_sum = sum(col_value(r, col_key) for r in data_rows)
        col_sum = round(col_sum, 2)
        summary_val = col_value(summary_row, col_key)
        diff = round(col_sum - summary_val, 2)
        passed = abs(diff) <= tolerance
        checks.append({
            "kind": "column_sum",
            "name": f"{col_label(col_key)} 列加总",
            "column_label": col_label(col_key),
            "col_sum": col_sum,
            "summary": summary_val,
            "diff": diff,
            "passed": passed,
            "detail": "" if passed else f"列加总 {col_sum:.2f} 与合计行 {summary_val:.2f} 相差 {diff:+.2f}"
        })

    # === B+C. 横向公式 ===
    for formula in validation_cfg.get("row_formulas", []) or []:
        name = formula.get("name", "<未命名公式>")
        lhs = formula.get("lhs")
        rhs_plus = formula.get("rhs_plus", []) or []
        rhs_minus = formula.get("rhs_minus", []) or []
        if not lhs:
            continue
        require_col(lhs, f"row_formulas[{name}].lhs")
        for k in rhs_plus:
            require_col(k, f"row_formulas[{name}].rhs_plus")
        for k in rhs_minus:
            require_col(k, f"row_formulas[{name}].rhs_minus")

        def rhs_of(row):
            return round(
                sum(col_value(row, k) for k in rhs_plus)
                - sum(col_value(row, k) for k in rhs_minus),
                2
            )

        # B: 合计行
        L = col_value(summary_row, lhs)
        R = rhs_of(summary_row)
        diff = round(L - R, 2)
        passed = abs(diff) <= tolerance
        checks.append({
            "kind": "row_formula_summary",
            "name": f"{name} (合计行)",
            "lhs_value": L,
            "rhs_value": R,
            "diff": diff,
            "passed": passed,
            "detail": "" if passed else f"合计行 lhs={L:.2f}，rhs={R:.2f}，差 {diff:+.2f}"
        })

        # C: 每个数据行
        failed_rows = []
        for ridx, r in enumerate(data_rows):
            L_r = col_value(r, lhs)
            R_r = rhs_of(r)
            d = round(L_r - R_r, 2)
            if abs(d) > tolerance:
                failed_rows.append({
                    "row_label": _row_label(r, f"第 {ridx + 1} 行"),
                    "lhs": L_r,
                    "rhs": R_r,
                    "diff": d,
                })
        total_rows = len(data_rows)
        passed_rows = total_rows - len(failed_rows)
        passed = len(failed_rows) == 0
        if passed:
            detail = f"{passed_rows}/{total_rows} 行通过"
        else:
            sample = "; ".join(
                f"{f['row_label']} 差 {f['diff']:+.2f}" for f in failed_rows[:3]
            )
            more = f"...（共 {len(failed_rows)} 行不通过）" if len(failed_rows) > 3 else ""
            detail = f"{passed_rows}/{total_rows} 行通过；失败例: {sample}{more}"
        checks.append({
            "kind": "row_formula_rows",
            "name": f"{name} (每行)",
            "total_rows": total_rows,
            "passed_rows": passed_rows,
            "failed_rows": failed_rows,
            "passed": passed,
            "detail": detail
        })

    passed_count = sum(1 for c in checks if c["passed"])
    failed_count = len(checks) - passed_count
    return {
        "ok": failed_count == 0,
        "passed_count": passed_count,
        "failed_count": failed_count,
        "checks": checks,
    }


def append_validation_sheet(file_bytes, validation_result,
                            sheet_name="验证结果",
                            source_filename=""):
    """
    在 Excel 字节流末尾追加一个 sheet，写入校验结果。原表数据不动。
    如果同名 sheet 已存在（重复处理场景），先删后建，保证幂等。
    返回新的字节流。

    注：.xls (Excel 97-2003) 格式为只读，不支持追加 sheet，
    此时返回原 bytes，由调用方提示用户。
    """
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = openpyxl.load_workbook(BytesIO(file_bytes))
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    meta_font = Font(italic=True, color="666666")
    pass_fill = PatternFill("solid", fgColor="E6F4EA")  # 淡绿
    fail_fill = PatternFill("solid", fgColor="FCE8E6")  # 淡红
    header_fill = PatternFill("solid", fgColor="EFEFEF")
    center = Alignment(horizontal="center", vertical="center")

    # 标题
    ws["A1"] = "工资表校验结果"
    ws["A1"].font = title_font
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = center

    # 元数据
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    passed = validation_result.get("passed_count", 0)
    failed = validation_result.get("failed_count", 0)
    meta_rows = [
        ("生成时间", now_str),
        ("源文件", source_filename or "(未提供)"),
        ("汇总", f"{passed} 项通过 / {failed} 项失败"),
    ]
    for i, (k, v) in enumerate(meta_rows, start=2):
        ws.cell(row=i, column=1, value=k).font = meta_font
        ws.cell(row=i, column=2, value=v).font = meta_font

    # 明细表头
    head_row = 2 + len(meta_rows) + 1  # 留一行空
    headers = ["校验项", "类型", "结果", "说明"]
    for cidx, h in enumerate(headers, start=1):
        c = ws.cell(row=head_row, column=cidx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    kind_label = {
        "column_sum": "纵向加总",
        "row_formula_summary": "横向公式",
        "row_formula_rows": "横向公式",
    }

    # 明细行
    for i, ck in enumerate(validation_result.get("checks", []), start=head_row + 1):
        ws.cell(row=i, column=1, value=ck.get("name", ""))
        ws.cell(row=i, column=2, value=kind_label.get(ck.get("kind"), ck.get("kind", "")))
        ws.cell(row=i, column=3, value="✅ 通过" if ck.get("passed") else "❌ 失败")
        ws.cell(row=i, column=4, value=ck.get("detail", ""))
        fill = pass_fill if ck.get("passed") else fail_fill
        for cidx in range(1, 5):
            ws.cell(row=i, column=cidx).fill = fill

    # 列宽
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 60

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def build_summary_workbook(parsed_list, validation_results, tf_columns,
                           write_back_sheet_name="验证结果"):
    """
    生成「工资发放汇总表」xlsx，包含两个 sheet：
      1) 汇总数据：每行一个附件，列同数据预览（文件名、年月、报表字段...、验证结果）
      2) 验证明细：所有附件的校验结果合并，多一列"附件名"区分

    解决两个问题：
      - .xls 附件无法在原文件内回写验证结果
      - 多附件场景下没有统一的"全局视图"
    """
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = openpyxl.Workbook()

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    meta_font = Font(italic=True, color="666666")
    header_fill = PatternFill("solid", fgColor="EFEFEF")
    pass_fill = PatternFill("solid", fgColor="E6F4EA")
    fail_fill = PatternFill("solid", fgColor="FCE8E6")
    center = Alignment(horizontal="center", vertical="center")
    money_font = Font(name="Consolas")

    # ===== Sheet 1: 汇总数据 =====
    ws = wb.active
    ws.title = "汇总数据"

    headers = ["文件名", "年月"] + [c["label"] for c in tf_columns] + ["验证结果"]
    ws["A1"] = "工资发放汇总数据"
    ws["A1"].font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws["A1"].alignment = center

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A2"] = f"生成时间：{now_str}    附件数：{len(parsed_list)}"
    ws["A2"].font = meta_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    head_row = 4
    for cidx, h in enumerate(headers, start=1):
        c = ws.cell(row=head_row, column=cidx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    for i, p in enumerate(parsed_list, start=head_row + 1):
        ws.cell(row=i, column=1, value=p.get("filename", ""))
        ws.cell(row=i, column=2, value=p.get("year_month", ""))
        for cidx, col in enumerate(tf_columns, start=3):
            v = p.get(col["key"], "")
            # 金额列尝试转 float，方便 Excel 求和/排序
            if isinstance(v, str):
                try:
                    v_num = float(v)
                    cell = ws.cell(row=i, column=cidx, value=v_num)
                    cell.number_format = "#,##0.00"
                    cell.font = money_font
                except ValueError:
                    ws.cell(row=i, column=cidx, value=v)
            else:
                ws.cell(row=i, column=cidx, value=v)
        # 验证结果列
        vr = validation_results.get(p.get("filename"))
        if vr is None:
            status = "未启用"
            fill = None
        elif vr["ok"]:
            status = f"✅ 全部通过 ({vr['passed_count']} 项)"
            fill = pass_fill
        else:
            status = f"⚠️ {vr['failed_count']} 项未通过 (共 {vr['passed_count']+vr['failed_count']} 项)"
            fill = fail_fill
        status_cell = ws.cell(row=i, column=len(headers), value=status)
        if fill is not None:
            status_cell.fill = fill

    # 末尾追加一行"总计"，对所有金额列求和（仅数值列）
    if parsed_list:
        total_row = head_row + 1 + len(parsed_list)
        ws.cell(row=total_row, column=1, value="总计").font = header_font
        for cidx, col in enumerate(tf_columns, start=3):
            # 试求和：把每行该列取出，能转 float 的相加
            total = 0.0
            has_num = False
            for p in parsed_list:
                v = p.get(col["key"], "")
                try:
                    total += float(v)
                    has_num = True
                except (ValueError, TypeError):
                    pass
            if has_num:
                cell = ws.cell(row=total_row, column=cidx, value=round(total, 2))
                cell.number_format = "#,##0.00"
                cell.font = Font(name="Consolas", bold=True)
                cell.fill = header_fill
            else:
                ws.cell(row=total_row, column=cidx, value="").fill = header_fill
        ws.cell(row=total_row, column=2, value="").fill = header_fill  # 年月列留空
        ws.cell(row=total_row, column=1).fill = header_fill
        ws.cell(row=total_row, column=len(headers), value="").fill = header_fill

    # 列宽
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    for ci in range(3, 3 + len(tf_columns)):
        ws.column_dimensions[ws.cell(row=head_row, column=ci).column_letter].width = 22
    ws.column_dimensions[ws.cell(row=head_row, column=len(headers)).column_letter].width = 36

    # ===== Sheet 2: 验证明细 =====
    ws2 = wb.create_sheet("验证明细")
    ws2["A1"] = "各附件验证明细"
    ws2["A1"].font = title_font
    ws2.merge_cells("A1:E1")
    ws2["A1"].alignment = center

    detail_headers = ["附件名", "校验项", "类型", "结果", "说明"]
    for cidx, h in enumerate(detail_headers, start=1):
        c = ws2.cell(row=3, column=cidx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    kind_label = {
        "column_sum": "纵向加总",
        "row_formula_summary": "横向公式",
        "row_formula_rows": "横向公式",
    }
    cur = 4
    for p in parsed_list:
        fn = p.get("filename", "")
        vr = validation_results.get(fn)
        if vr is None:
            ws2.cell(row=cur, column=1, value=fn)
            ws2.cell(row=cur, column=2, value="(未启用校验)").fill = header_fill
            cur += 1
            continue
        for ck in vr.get("checks", []):
            ws2.cell(row=cur, column=1, value=fn)
            ws2.cell(row=cur, column=2, value=ck.get("name", ""))
            ws2.cell(row=cur, column=3, value=kind_label.get(ck.get("kind"), ck.get("kind", "")))
            ws2.cell(row=cur, column=4, value="✅ 通过" if ck.get("passed") else "❌ 失败")
            ws2.cell(row=cur, column=5, value=ck.get("detail", ""))
            fill = pass_fill if ck.get("passed") else fail_fill
            for cidx in range(1, 6):
                ws2.cell(row=cur, column=cidx).fill = fill
            cur += 1

    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 38
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 10
    ws2.column_dimensions["E"].width = 60

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def generate_title(unit_names, year_month, amounts):
    """
    unit_names: list of unit names (ordered by amount desc)
    amounts: list of transfer amounts (same order)
    year_month: "2026年03月" format
    """
    n = len(unit_names)
    if n == 0:
        return f"{year_month}工资发放请示"
    if n == 1:
        return f"{unit_names[0]}{year_month}工资发放请示"
    if n == 2:
        return f"{unit_names[0]}、{unit_names[1]}{year_month}工资发放请示"
    # 3+ units
    return f"{unit_names[0]}、{unit_names[1]}等{n}家单位{year_month}工资发放请示"


def main():
    ui_config = CONFIG.get("ui", DEFAULT_CONFIG["ui"])
    template_name = ui_config.get("template_name", "工资发放审批")
    description = ui_config.get("description", "")

    st.title(f"📋 {template_name}")
    if description:
        st.info(description)

    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False
    if "user_id" not in st.session_state:
        st.session_state.user_id = ""
    if "union_id" not in st.session_state:
        st.session_state.union_id = ""
    if "dept_id" not in st.session_state:
        st.session_state.dept_id = 0
    if "user_name" not in st.session_state:
        st.session_state.user_name = ""

    client = DingTalkClient(APP_KEY, APP_SECRET, AGENT_ID, PROCESS_CODE)

    # Section 1: Phone Login
    if not st.session_state.logged_in:
        st.subheader("手机号登录")
        mobile = st.text_input("手机号", value="")
        if st.button("登录"):
            if not mobile:
                st.error("请输入手机号")
                return
            try:
                user_id = client.get_user_by_mobile(mobile)
                info = client.get_user_info(user_id)
                st.session_state.logged_in = True
                st.session_state.user_id = user_id
                st.session_state.union_id = info["unionId"]
                st.session_state.dept_id = info["deptId"]
                st.session_state.user_name = info["name"]
                st.success(f"登录成功：{info['name']} ({user_id})")
                st.rerun()
            except Exception as e:
                st.error(f"登录失败：{e}")
        return

    # Section 2: File Upload
    st.subheader(f"欢迎，{st.session_state.user_name}")
    col1, col2 = st.columns([1, 1])
    with col2:
        if st.button("退出登录"):
            for key in ["logged_in", "user_id", "union_id", "dept_id", "user_name"]:
                st.session_state.pop(key, None)
            st.rerun()
    uploaded_files = st.file_uploader(
        "上传工资表", type=["xlsx", "xls"], accept_multiple_files=True
    )

    if not uploaded_files:
        st.info("请上传一个或多个 Excel 工资表文件")
        return

    # Parse each file
    parsed_list = []
    for upfile in uploaded_files:
        file_bytes = upfile.read()
        upfile.seek(0)
        parsed = parse_excel(file_bytes, upfile.name)
        if parsed:
            parsed_list.append({"filename": upfile.name, **parsed})

    if not parsed_list:
        st.error("未能解析任何文件，请检查格式")
        return

    # 标题 vs 文件名年月一致性检查（取标题为准；不一致时提醒制表人核对）
    for p in parsed_list:
        t = p.get("year_month_from_title", "")
        f = p.get("year_month_from_filename", "")
        if t and f and t != f:
            st.warning(
                f"⚠️ {p['filename']}：标题中年月「{t}」与文件名年月「{f}」不一致，"
                f"已以**标题**为准。请确认报表标题是否需要更正。"
            )
        elif not t and f:
            st.warning(
                f"⚠️ {p['filename']}：报表标题中未识别到年月，已退回使用文件名年月「{f}」。"
                f"建议在标题中明确写出年月。"
            )

    # 工资表内容校验
    val_cfg = CONFIG.get("validation", DEFAULT_CONFIG["validation"])
    validation_results = {}  # {filename: validation_result}
    submit_blocked = False
    if val_cfg.get("enabled"):
        excel_cols_def = CONFIG.get("excel", {}).get("columns", {})
        for p in parsed_list:
            try:
                vr = validate_payroll(p, val_cfg, excel_cols_def)
            except ValueError as e:
                st.error(f"⚠️ {p['filename']} 校验配置错误：{e}")
                vr = None
                if val_cfg.get("strict"):
                    submit_blocked = True
            validation_results[p["filename"]] = vr
            if vr is not None and not vr["ok"] and val_cfg.get("strict"):
                submit_blocked = True

    # Preview table
    st.subheader("数据预览")
    preview_data = []
    tf_columns = CONFIG["table_field"]["columns"]
    for p in parsed_list:
        row = {"文件名": p["filename"], "年月": p["year_month"]}
        for col in tf_columns:
            row[col["label"]] = p[col["key"]]
        # 验证结果列
        if val_cfg.get("enabled"):
            vr = validation_results.get(p["filename"])
            if vr is None:
                row["验证结果"] = "⚠️ 配置错误"
            elif vr["ok"]:
                row["验证结果"] = f"✅ 全部通过 ({vr['passed_count']} 项)"
            else:
                row["验证结果"] = f"⚠️ {vr['failed_count']}/{vr['passed_count'] + vr['failed_count']} 项未通过"
        preview_data.append(row)
    st.dataframe(preview_data)

    # 生成审批标题（提前到这里，汇总表文件名要从标题派生）
    # Sort by transfer_total desc
    sorted_items = sorted(
        parsed_list,
        key=lambda x: float(x["transfer_total"]) if x["transfer_total"] else 0,
        reverse=True,
    )
    unit_names = [p["unit_name"] for p in sorted_items]
    amounts = [p["transfer_total"] for p in sorted_items]
    year_month = sorted_items[0]["year_month"] if sorted_items else ""
    title = generate_title(unit_names, year_month, amounts)

    # 生成「工资发放汇总表」xlsx：把数据预览 + 所有附件的验证明细打包成一个文件
    # 用途：(1) .xls 附件无法回写验证 sheet，靠这里集中展示；
    #      (2) 多附件场景下提供全局视图；(3) 作为额外附件随审批一起归档
    # 文件名从审批标题派生：把"工资发放请示"换成"工资发放汇总表"，与审批主题保持一致
    if "工资发放请示" in title:
        summary_filename = title.replace("工资发放请示", "工资发放汇总表") + ".xlsx"
    else:
        summary_filename = (
            f"{title}汇总表.xlsx" if title else "工资发放汇总表.xlsx"
        )
    try:
        summary_bytes = build_summary_workbook(
            parsed_list,
            validation_results,
            tf_columns,
            write_back_sheet_name=val_cfg.get("write_back_sheet_name", "验证结果"),
        )
        st.download_button(
            label=f"📥 下载汇总表（{summary_filename}）",
            data=summary_bytes,
            file_name=summary_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        summary_bytes = None
        st.warning(f"⚠️ 生成汇总表失败：{e}")

    # 提示：若有失败，告诉用户去看 Excel sheet
    if val_cfg.get("enabled") and any(
        vr is not None and not vr["ok"] for vr in validation_results.values()
    ):
        st.info(
            f"📋 详细校验明细已写入附件的「{val_cfg.get('write_back_sheet_name', '验证结果')}」sheet（仅 .xlsx 附件），"
            f"或参见上方汇总表的「验证明细」sheet。"
        )

    # 显示审批标题（title 已在上方汇总表前生成）
    st.text_input("审批标题（自动生成）", value=title, disabled=True)

    # Submit button
    if submit_blocked:
        st.error("⚠️ 校验未通过且当前为严格模式，请修改 Excel 后重新上传")
    if st.button("提交审批", disabled=submit_blocked):
        progress_bar = st.progress(0)
        status_text = st.empty()

        try:
            user_id = st.session_state.user_id
            union_id = st.session_state.union_id
            dept_id = st.session_state.dept_id

            # Upload each file to DingTalk
            attachments = []
            # 总步数 = 每个原附件 + 汇总表（如果生成成功） + 创建审批实例
            total_steps = len(uploaded_files) + (1 if summary_bytes else 0) + 1
            done_steps = 0
            for i, upfile in enumerate(uploaded_files):
                status_text.text(f"正在上传：{upfile.name} ...")
                file_bytes = upfile.read()
                upfile.seek(0)

                # 若启用了 write_back_sheet，把校验结果作为新 sheet 追加到 Excel 后再上传
                vr = validation_results.get(upfile.name)
                if (vr is not None
                    and val_cfg.get("enabled")
                    and val_cfg.get("write_back_sheet", True)
                    and not upfile.name.lower().endswith(".xls")):
                    file_bytes = append_validation_sheet(
                        file_bytes,
                        vr,
                        sheet_name=val_cfg.get("write_back_sheet_name", "验证结果"),
                        source_filename=upfile.name,
                    )

                # Authorize before EACH upload
                space_id = client.authorize_upload(user_id)
                result = client.upload_file(
                    file_bytes, upfile.name, union_id, space_id
                )
                attachments.append(result)
                done_steps += 1
                progress_bar.progress(done_steps / total_steps)

            # 把汇总表也作为额外附件上传（如果生成成功）
            if summary_bytes:
                status_text.text(f"正在上传：{summary_filename} ...")
                space_id = client.authorize_upload(user_id)
                result = client.upload_file(
                    summary_bytes, summary_filename, union_id, space_id
                )
                attachments.append(result)
                done_steps += 1
                progress_bar.progress(done_steps / total_steps)

            # Build table rows
            table_rows = []
            tf_columns = CONFIG["table_field"]["columns"]
            for p in parsed_list:
                table_rows.append(
                    [
                        {"name": col["label"], "value": p[col["key"]]}
                        for col in tf_columns
                    ]
                )

            status_text.text("正在创建审批实例...")
            instance_id = client.create_approval(
                user_id=user_id,
                dept_id=dept_id,
                title=title,
                attachments=attachments,
                table_rows=table_rows,
                note="",
            )
            progress_bar.progress(1.0)
            status_text.empty()

            if instance_id:
                st.success(f"审批创建成功！instanceId：{instance_id}")
            else:
                st.warning("审批创建完成，但未返回 instanceId")
        except Exception as e:
            progress_bar.empty()
            status_text.empty()
            st.error(f"提交失败：{e}")


if __name__ == "__main__":
    main()
